import os
import sys
import json
import time
import hashlib
import openpyxl
import warnings
import traceback


warnings.filterwarnings('ignore')

if not os.path.exists('setting.py'):
    with open('setting.py', 'w') as fl:
        fl.write("FILE_SAVE_PATH = r''\n")
        fl.write("NEED_COMBINE_DICTORY_PATH = r''\n")
        fl.write("MODE = 1\n")
    print('''
因第一次运行，请先填好 setting.py 文件。
可以使用记事本打开。请在引号中填写字段。

FILE_SAVE_PATH 填写合并完后文件保存路径；

NEED_COMBINE_DICTORY_PATH 填写需要合并的文件夹路径；
如果都不填写，则默认合并脚本路径下的后缀名为 .xlsx 的
文件，并且生成的文件也保存在该路径下。后续生成的文件不
会进入合并的列表中。

MODE 表示模式选择，目前只有 0 和 1 两种模式：
模式 0 : 普通合并所有 Excel 文件；
模式 1 : 合并所有 Excel 文件，从第二个文件开始不合并
         文件的第一行，因为有很多情况下这些 Excel 文
         件都是同样形式；

配置完成后，请重新运行本程序。
'''
          )
    os.system('pause')
    sys.exit()

from setting import *

if MODE not in (0, 1):
    print('请在 setting.py 中设置正确的 MODE 值，')
    print('目前只支持 0 和 1,现在程序停止。')
    os.system('pause')
    sys.exit()


def md5_text(text):
    md5 = hashlib.md5()
    md5.update(text.encode())
    return md5.hexdigest()

def safe_remove(path):
    if os.path.exists(path):
        os.remove(path)


class JsonDb(dict):

    @classmethod
    def from_file(cls, path):
        with open(path, 'r', encoding='utf-8')as fl:
            data = json.load(fl)
        return cls(data)

    def save(self, path):
        with open(path, 'w', encoding='utf-8')as fl:
            json.dump(self, fl, ensure_ascii=False, indent=4)


class CombineExcelFiles:

    def __init__(self):

        self.filename = '合并完成.xlsx'
        self.log_file = LOG_FILE
        self.save_file = os.path.join(FILE_SAVE_PATH, self.filename)
        self.combine_file_path = NEED_COMBINE_DICTORY_PATH

    def run(self, mode=MODE):
        try:
            if not os.path.exists(self.save_file) or not os.path.exists(self.log_file):
                safe_remove(self.save_file)
                safe_remove(self.log_file)
                self.log_flag = False
                self.log_data = JsonDb()
                self.log_data['path_md5'] = []
                self.create_excel()
                self.combine(mode)
            else:
                print('正在加载初始表格...')
                self.load_excel(self.save_file)
                print('初始表格加载完成。')
                self.log_flag = True
                self.log_data = JsonDb.from_file(self.log_file)
                self.combine(mode)

        except Exception as e:
            traceback.print_exc()
            os.system('pause')
        finally:
            self.close()

    def create_excel(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.max_row = 0
        self.max_column = 0

    def load_excel(self, path):
        self.wb = openpyxl.load_workbook(path)
        self.ws = self.wb.active
        self.max_row = self.ws.max_row
        self.max_column = self.ws.max_column

    def load_and_combine(self, excel_abs_path, mode=MODE):

        wb = openpyxl.load_workbook(excel_abs_path)
        ws = wb.active
        # 得到目标表格最大的行列数
        max_row = ws.max_row
        max_column = ws.max_column
        if mode == 0:
            for row in range(1, max_row + 1):
                for column in range(1, max_column + 1):
                    self.ws.cell(
                        row=self.max_row + 1, column=column).value = ws.cell(row=row, column=column).value
                self.max_row += 1
        elif mode == 1:
            if self.max_row == 0:
                for row in range(1, max_row + 1):
                    for column in range(1, max_column + 1):
                        self.ws.cell(
                            row=self.max_row + 1, column=column).value = ws.cell(row=row, column=column).value
                    self.max_row += 1
            else:
                for row in range(2, max_row + 1):
                    for column in range(1, max_column + 1):
                        self.ws.cell(
                            row=self.max_row + 1, column=column).value = ws.cell(row=row, column=column).value
                    self.max_row += 1
        else:
            pass
        wb.close()

    def combine(self, mode=MODE):

        # 不同的 mode 代表着不同的表格合并策略
        # 默认 0 代表着全合并
        # 数字 1 代表着除了第一个表格，其他表格忽略掉第一行
        target_file_list = os.listdir(self.combine_file_path)
        for root, dirs, files in os.walk(self.combine_file_path):
            for file in files:
                if file.endswith('.xlsx') and '合并' not in file:
                    excel_abs_path = os.path.join(root, file)
                    excel_abs_path_md5 = md5_text(excel_abs_path)
                    if excel_abs_path_md5 not in self.log_data['path_md5']:
                        print(excel_abs_path)
                        self.load_and_combine(excel_abs_path)
                        self.log_data['path_md5'].append(excel_abs_path_md5)

    def close(self):
        try:
            self.wb.save(self.save_file)
            self.log_data.save(self.log_file)
        except PermissionError:
            print('保存失败！！！\n')
            print(f'请先关闭[{self.save_file}]才能正常保存。')
        finally:
            self.wb.close()

if __name__ == '__main__':
    print('提示：只有后缀名为 xlsx 的才能合并。')
    print('正在合并工作表...\n')

    start_time = time.time()
    c = CombineExcelFiles()
    c.run(mode=MODE)
    end_time = time.time()

    print('\n合并完成。共耗时 {} 秒。\n'.format(end_time - start_time))
    print('目标文件位置为：')
    print(c.save_file)
    print()
    os.system('pause')
